"""Excel lead importer."""

from pathlib import Path

import structlog
from openpyxl import Workbook, load_workbook

from src.core.db import DEFAULT_DB_PATH, insert_lead

log = structlog.get_logger()


def import_leads(excel_path: Path, db_path: Path = DEFAULT_DB_PATH) -> dict:
    """Import leads from Excel file.

    Expected columns: email, first_name, last_name, company, title, linkedin_url

    Returns dict with imported and skipped counts.
    """
    wb = load_workbook(excel_path)
    ws = wb.active

    # Get header row
    headers = [cell.value.lower().strip() if cell.value else "" for cell in ws[1]]

    required = {"email", "first_name"}
    if not required.issubset(set(headers)):
        raise ValueError(f"Excel must have columns: {required}. Found: {headers}")

    # Map column indices
    col_map = {name: idx for idx, name in enumerate(headers)}

    imported = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        email = row[col_map["email"]]
        if not email:
            continue

        email = email.strip().lower()
        first_name = row[col_map["first_name"]]

        if not first_name:
            log.warning("skipping_row_no_first_name", email=email)
            skipped += 1
            continue

        lead_id = insert_lead(
            db_path=db_path,
            email=email,
            first_name=first_name.strip(),
            last_name=row[col_map.get("last_name", -1)] if col_map.get("last_name") is not None and col_map["last_name"] < len(row) else None,
            company=row[col_map.get("company", -1)] if col_map.get("company") is not None and col_map["company"] < len(row) else None,
            title=row[col_map.get("title", -1)] if col_map.get("title") is not None and col_map["title"] < len(row) else None,
            linkedin_url=row[col_map.get("linkedin_url", -1)] if col_map.get("linkedin_url") is not None and col_map["linkedin_url"] < len(row) else None,
        )

        if lead_id:
            log.info("lead_imported", email=email, lead_id=lead_id)
            imported += 1
        else:
            log.info("lead_skipped_duplicate", email=email)
            skipped += 1

    return {"imported": imported, "skipped": skipped}


def create_example_excel(output_path: Path) -> None:
    """Create an example Excel file showing expected format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Headers
    ws.append(["email", "first_name", "last_name", "company", "title", "linkedin_url"])

    # Example rows
    ws.append([
        "sarah@glossybrand.com",
        "Sarah",
        "Chen",
        "Glossy Brand",
        "Marketing Director",
        "https://linkedin.com/in/sarahchen"
    ])
    ws.append([
        "mike@acmeco.com",
        "Mike",
        "Johnson",
        "Acme Co",
        "Head of Growth",
        "https://linkedin.com/in/mikej"
    ])

    wb.save(output_path)
